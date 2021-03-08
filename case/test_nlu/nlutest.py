# -*- coding: utf-8 -*-
import concurrent.futures
import json
import requests
from copy import deepcopy
from openpyxl import load_workbook
from base.operatexcel import operatexcel
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
import yaml
import os
from base.log import logger
from os import path
import uuid
from base.readpath import getpath
import time

d = path.dirname(__file__)
parent_path1 = os.path.dirname(d)
parent_path2 = os.path.dirname(parent_path1)


def runnlu(cases):
    pid = cases["pid"]
    newnlu = cases["newnlu"].lower()
    print("查询语料： {}".format(cases["query"]))
    if newnlu == "true":
        try:
            expectintentlist = eval(cases["expectintents"])
            for k,v in expectintentlist.items():
                expectintentlist[k]=expectintentlist[k].strip().split(",")
        except Exception as e:
            print("expectintents 转json失败： ", e)
            expectintentlist = ""
        if isinstance(expectintentlist, dict):
            pass
        else:
            expectintentlist = {}
            # expectintentlist = [intent.strip() for intent in cases["expectintents"].replace("\"", "").split(",")]
            expectprointentlist = []
            expectglbintentlist = []
            expectsimintentlist = []
            for intents in cases["expectintents"].replace("\\n", "").replace("'", "").replace("{", ""). \
                    replace("}", "").replace("\"", "").replace(" ", "").strip().split(";"):
                furintent = intents.split(":")
                if furintent[0] == "PRO":
                    expectprointentlist = furintent[1].split(",")
                if furintent[0] == "GLB":
                    expectglbintentlist = furintent[1].split(",")
                if furintent[0] == "SIM":
                    expectsimintentlist = furintent[1].split(",")

            expectintentlist["PRO"] = expectprointentlist
            expectintentlist["GLB"] = expectglbintentlist
            expectintentlist["SIM"] = expectsimintentlist

        print("expectintentlist type is: " + str(type(expectintentlist)))
        print("expectintentlist is: " + json.dumps(expectintentlist, ensure_ascii=False))
        prointentlist = []
        glbintentlist = []
        simintentlist = []
        for intents in cases["queryintent"].replace("\n", "").replace("\"", "").replace(" ", "").split(";"):
            furintent = intents.split(":", 1)
            if furintent[0] == "PRO":
                prointentlist = furintent[1].split(",")
            if furintent[0] == "GLB":
                glbintentlist = furintent[1].split(",")
            if furintent[0] == "SIM":
                sims = furintent[1].split(",")
                simintentdict = {}
                for sim in sims:
                    intent_threshold = sim.split(":")
                    simintentdict["intentName"] = intent_threshold[0]
                    if len(intent_threshold) > 1:
                        simintentdict["threshold"] = float(intent_threshold[1])
                    else:
                        simintentdict["threshold"] = 0.6  # 为每个相似意图设置默认阈值
                    simintentlist.append(deepcopy(simintentdict))
                    simintentdict.clear()

        r = NluTest.nlufusionquery(pid, cases["query"], cases["env"], newnlu, proinentlist=prointentlist,
                                   glbintentlist=glbintentlist, simintentlist=simintentlist)
        print("查询语料： {}".format(cases["query"]))
        cases["reintent"] = r[0]
        cases["reslots"] = r[1]
        # expectintents=queryexpectintent
        reintent = r[0]
        if type(reintent) is dict:
            cases["testre_intent"] = True
            for k,v in expectintentlist.items():
                if len(expectintentlist[k])>0:
                    reintents=reintent[k].split(",")
                    for intent in v:
                        if not reintents.__contains__(intent):
                            cases["testre_intent"] = False

            # for k, v in reintent.items():
            #     for intent in reintent[k].split(","):
            #         if intent == "":
            #             pass
            #         else:
            #             if expectintentlist[k].__contains__(intent):
            #                 cases["testre_intent"] = True
            #                 # pass
            #             else:
            #                 # flag = False
            #                 cases["testre_intent"] = False

        if cases["expectslots"] is None or cases["expectslots"].strip() == "":
            cases["testre_slots"] = None
        else:
            testre_slots = True
            try:
                for intent, slots in eval(cases["expectslots"]).items():
                    reslots = r[1][intent]
                    slotname = {}
                    for slot in slots:
                        slotname[slot["name"]] = False
                    for slot in slots:
                        # print("slot: "+str(slot))
                        for reslot in reslots:
                            # print("reslot: "+str(reslot))
                            if reslot["name"] == slot["name"] and reslot["value"] == slot["value"]:
                                slotname[slot["name"]] = True
                                break
                            # else:
                            #     testre_slots = False
                    for k, v in slotname.items():
                        if v is False:
                            testre_slots = False
                    slotname.clear()
                cases["testre_slots"] = testre_slots
            except Exception as e:
                print("槽位对比异常默认失败： ", e)
                testre_slots = False
                cases["testre_slots"] = testre_slots
    elif newnlu == "ccdmnlu":
        expectintentlist = [intent.strip() for intent in cases["expectintents"].replace("\"", "").split(",")]
        r = NluTest.nlufusionquery(pid, cases["query"], cases["env"], newnlu, intents=None)
        cases["reintent"] = r[0]
        cases["reslots"] = r[1]
        if sorted(r[0].split(",")) == sorted(expectintentlist):
            cases["testre_intent"] = True
        else:
            cases["testre_intent"] = False

        if cases["expectslots"] is None:
            cases["testre_slots"] = None
        else:
            testre_slots = False
            try:
                for intent, slots in eval(cases["expectslots"]).items():
                    reslots = r[1][intent]
                    for slot in slots:
                        for reslot in reslots:
                            if reslot["slot"] == slot["slot"] and reslot["value"] == slot["value"]:
                                testre_slots = True
                            else:
                                testre_slots = False
                cases["testre_slots"] = testre_slots
            except Exception as e:
                print("预期槽位解析json失败，使用默认： ", e)
                testre_slots = False
                cases["testre_slots"] = testre_slots
    else:
        queryintent = cases["queryintent"].replace("\"", "").split(",")
        queryintentlist = [intent.strip() for intent in queryintent]
        expectintentlist = [intent.strip() for intent in cases["expectintents"].replace("\"", "").split(",")]
        r = NluTest.nlufusionquery(pid, cases["query"], cases["env"], newnlu, intents=queryintentlist)
        print("查询语料： {}".format(cases["query"]))
        cases["reintent"] = r[0]
        cases["reslots"] = r[1]
        # expectintents=queryexpectintent

        # for intent in r[0].split(","):
        #     if expectintentlist.__contains__(intent):
        #         cases["testre_intent"] = True
        #         # pass
        #     else:
        #         # flag = False
        #         cases["testre_intent"] = False
        if sorted(r[0].split(",")) == sorted(expectintentlist):
            cases["testre_intent"] = True
        else:
            cases["testre_intent"] = False

        if cases["expectslots"] is None:
            cases["testre_slots"] = None
        else:
            testre_slots = False
            try:
                for intent, slots in eval(cases["expectslots"]).items():
                    reslots = r[1][intent]
                    for slot in slots:
                        for reslot in reslots:
                            if reslot["slot"] == slot["slot"] and reslot["value"] == slot["value"]:
                                testre_slots = True
                            else:
                                testre_slots = False
                cases["testre_slots"] = testre_slots
            except Exception as e:
                print(e)
                testre_slots = False
                cases["testre_slots"] = testre_slots
    if cases["testre_intent"] is True and cases["testre_slots"] is True:
        cases["testre"] = True
        flag = True
    elif cases["testre_intent"] is True and cases["testre_slots"] is None:
        cases["testre"] = True
        flag = True
    else:
        cases["testre"] = False
        flag = False
    print("casere is: " + json.dumps(cases, ensure_ascii=False))
    return flag, cases


def getexcelsheetconfig(filename, sheetnames=None):
    envcell = "F1"
    wb = load_workbook(filename)
    envvalue = ["test", "beta", "alpha", "prod"]
    if sheetnames is None:
        sheetnames_waitcollect = wb.sheetnames
    else:
        sheetnames_waitcollect = sheetnames
    env_sheet = {}
    for sheet in sheetnames_waitcollect:
        if wb[sheet][envcell].value is not None and wb[sheet][envcell].value.lower() in envvalue:
            if env_sheet.get(wb[sheet][envcell].value.lower()) is None:
                env_sheet[wb[sheet][envcell].value.lower()] = []
            env_sheet[wb[sheet][envcell].value.lower()].append(sheet)
    excelconfig = []
    for env, sheets in env_sheet.items():
        excelconfig.append({"file": filename, "env": env, "sheetnames": sheets})
    return excelconfig


def getexcelconfig(casename, sheetnames=None):
    if os.path.exists(casename):
        print("自定义路径文件存在")
        config = getexcelsheetconfig(casename, sheetnames)
    else:
        rootdir = getpath()
        file = rootdir + "/data/test_json/" + casename
        print(file)
        if os.path.exists(file):
            print("文件存在: ", file)
            config = getexcelsheetconfig(file, sheetnames)
    return config


def getcases(filename, sheetnams=None):
    caseconfs_original = NluTest.readyaml()
    caseconfs = []
    if os.path.exists(filename):
        caseconfs = getexcelsheetconfig(filename, sheetnames=sheetnams)
    else:
        rootdir = getpath()
        file = rootdir + "/data/test_nlu/" + filename
        if os.path.exists(file):
            caseconfs = getexcelsheetconfig(file, sheetnames=sheetnams)

    # 注销yaml获取用例方式
    # if len(caseconfs) == 0:
    #     print("excel配置未获取用例为零，开始查询yaml配置")
    #     for casedata in caseconfs_original["filenames"]:
    #         if casedata["file"] == filename:
    #             caseconfs.append(casedata)
    #             break
    #     for caseconf in caseconfs:
    #         caseconf["file"] = parent_path2 + "/data/test_nlu/" + caseconf["file"]
    print("caseconfs: ", caseconfs)
    cases = []
    for caseconf in caseconfs:
        sheetname = caseconf["sheetnames"]
        file = caseconf["file"]
        print(file, sheetname)
        a = NluTest(file, sheetname)
        a.move_reintent()
        a.move_precision_rate()
        a.reset()
        a.save()
        case = a.main()
        for item in case:
            item["env"] = caseconf["env"]
        cases.extend(case)
        case.clear()
        print("执行完毕")
    if len(cases) == 0:
        raise RuntimeError("没有可运行用例，配置文件配置错误")
    return cases


def after_execute(re):
    files_name = {}
    for item in re:
        if files_name.get(item["filename"]) is None:
            files_name[item["filename"]] = set()
            files_name[item["filename"]].add(item["sheetname"])
        else:
            files_name[item["filename"]].add(item["sheetname"])
    filesobject = {}
    for k, v in files_name.items():
        filesobject[k] = NluTest(k, v)
    for r in re:
        filesobject[r["filename"]].parsere(r)
    for k, v in filesobject.items():
        v.casecompare()
        v.save()


def postnlu(cases, threadnum):
    """
    多线程执行 dm 请求
    :param threadnum:
    :param cases:
    :return:
    """
    with concurrent.futures.ThreadPoolExecutor(max_workers=threadnum) as executor:
        to_do = []
        for case in cases:
            future = executor.submit(runnlu, case)
            to_do.append(future)

        # for future in concurrent.futures.as_completed(to_do):
        #     # future.result()
        #     pass


def runcase(filename=None, threadnum=None, sheetnames=None):
    cases = getcases(filename, sheetnames)
    postnlu(cases, threadnum)
    after_execute(cases)
    # with open("case","a",encoding="utf-8")  as case_file:
    #     case_file.write(str(cases))
    #     case_file.write("\n")


class NluTest:
    excelclown = {
        "query": "A",
        "resultintents": "B",
        "expectintents": "C",
        "compareintent": "G",  # 意图对比
        "compareslot": "H",  # 槽位对比
        "comparedresult": "I",  # 对比结果列
        "caserow": 2,
        "precisionrate": "J",
        "repast_start": "L",
        "repast_end": "Q",
        "queryintet": "D",
        "expectslots": "F",
        "reexpectslots": "E",

    }
    config = {
        "productid": "B1",
        "nlutype": "D1"
    }
    aihivebox_url = "http://ba-bnlu-test.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7"

    def __init__(self, file, sheetnames):
        self.sheetnames = sheetnames
        self.file = file
        self.wb = load_workbook(self.file)

    # def nlutest(self):
    #     with open("case.yaml", "r", encoding="utf") as f:
    #         lines = f.readlines()
    #         for line in lines:
    #             a = self.nlufusionquery(line.strip())
    #             print("{}  response is {}".format(line.strip(), a))

    def query_ba_aihivebox(self, query):
        """
        query 意图查询语料，
        :param query:
        :return: str,意图名字 "name1,name2"
        """
        dataraw = {"params": {"request": {"refText": query, "coreType": "cn.dlg.ita", "res": "aihivebox",
                                          "env": "use_slot_index=1"}}}
        data = "params" + "=" + json.dumps(dataraw["params"], ensure_ascii=False)
        print(data)
        try:
            r = requests.post(url=self.aihivebox_url, data=data.encode("utf-8"))
            if r.status_code == 200:
                param = json.loads(r.text)["result"]["semantics"]["request"]["param"]
                rstr = ''
                for k, v in param.items():
                    rstr = rstr + str(k) + ":" + str(v) + ","
                    return rstr.strip(",")
            else:
                return "意图参数问题，返回非200"
            # print(r.text)
        except Exception as e:
            print("意图请求错误: ", e)
            return "意图请求错误"

    @staticmethod
    def nlufusionaics(query, env, proinentlist, pid, glbintentlist, simintentlist):
        url = {
            "test": "http://internal-test.talkinggenie.com/nlu-fusion/fusion/test/aics",
            "alpha": "http://internal-alpha.talkinggenie.com/nlu-fusion/fusion/test/aics",
            "beta": "http://internal-beta.talkinggenie.com/nlu-fusion/fusion/test/aics",
            "prod": " http://internalofficeread.talkinggenie.com/nlu-fusion/nlu-fusion/fusion/test/aics"
        }
        aicsurl = url[env]
        data = {
            "sentence": query,
            "PROIntentNameList": proinentlist,
            "SIMIntentNameList": simintentlist,
            "env": "sandbox",
            "projectId": pid,
            "GLBIntentNameList": glbintentlist
        }
        logger.info("request is {}".format(json.dumps(data, ensure_ascii=False)))
        r = requests.post(url=aicsurl, json=data)
        logger.info("resp : {}".format(r.text))
        try:
            # print(r.text)
            reintents = {}
            slotvalue = {}
            intentlist = r.json()["result"]["intentList"]
            if len(intentlist) > 0:
                # rintent = ""
                glbintent = ""
                prointent = ""
                simintent = ""
                for intent in intentlist:
                    scope = intent.get("scope")
                    if scope == "GLB" or scope is None:
                        # pass
                        glbintent = glbintent + "," + intent["name"]
                    if scope == "SIM":
                        simintent = simintent + "," + intent["name"]
                    else:
                        prointent = prointent + "," + intent["name"]
                    # rintent = rintent + "," + intent["intent_name"]
                glbintent = glbintent.strip(",")
                prointent = prointent.strip(",")
                simintent = simintent.strip(",")
                # reintents="PRO:"+prointent+";"+"GLB"+glbintent
                reintents["PRO"] = prointent
                reintents["GLB"] = glbintent
                reintents["SIM"] = simintent
            else:
                reintents["PRO"] = "空意图"
                reintents["GLB"] = "空意图"
                reintents["SIM"] = "空意图"
                # reintents = "空意图"
            if len(intentlist) > 0:
                for intent in intentlist:
                    # intentslot = {}
                    slotvalue[intent["name"]] = intent["slotList"]
            if reintents == "空意图":
                slotvalue = "无槽位"
            return reintents, slotvalue
        except Exception as e:
            print(e)
            return "请求失败:statuscode,{},response {}".format(r.status_code, r.text)

    @staticmethod
    def ccdmnlu(pid, query, env):
        ccdmnlu_url = {
            "dev": "http://nlu-dev.talkinggenie.com/ccdmnlu/yto/intent",
            "test": "http://nlu-test.talkinggenie.com/ccdmnlu/yto/intent",
            "alpha": "http://nlu-alpha.talkinggenie.com/ccdmnlu/yto/intent",
            "beta": "http://nlu-beta.talkinggenie.com/ccdmnlu/yto/intent"
        }
        ccdmnluurl = ccdmnlu_url[env]
        uustr = str(uuid.uuid1()).replace('-', '')
        data = {"senderId": uustr, "query": query,
                "context": {"expectUserIntent": []}, "pid": pid}
        logger.info("request is {}".format(json.dumps(data, ensure_ascii=False)))
        r = requests.post(url=ccdmnluurl, json=data)
        logger.info("resp : {}".format(r.text))
        try:
            slotvalue = {}
            intentlist = r.json()["intents"][0]["dialog_acts"]
            if len(intentlist) > 0:
                rintent = ""
                for intent in intentlist:
                    rintent = rintent + "," + intent["intent_name"]
                reintents = rintent.strip(",")
            else:
                reintents = "空意图"
            if len(intentlist) > 0:
                for intent in intentlist:
                    # intentslot = {}
                    slotvalue[intent["intent_name"]] = intent["slots"]
            if reintents == "空意图":
                slotvalue = "无槽位"
            return reintents, slotvalue
        except Exception as e:
            print(e)
            return "请求失败:statuscode,{},response {}".format(r.status_code, r.text)

    @staticmethod
    def nlufusioninent(pid, query, env, intents):
        nlufusionurl = {
            "test": "http://internal-test.talkinggenie.com/nlu-fusion/nlu/intent",
            "alpha": "http://internal-alpha.talkinggenie.com/nlu-fusion/nlu/intent",
            "beta": "http://internal-beta.talkinggenie.com/nlu-fusion/nlu/intent"
        }
        url = nlufusionurl[env]
        uustr = str(uuid.uuid1()).replace('-', '')
        data = {
            "intents": intents,
            "productId": pid,
            "query": query,
            "ranking": "model&rule-slot",
            "sessionId": "18800585287&" + uustr
        }
        # print("request is {}".format(json.dumps(data, ensure_ascii=False)))
        logger.info("request is {}".format(json.dumps(data, ensure_ascii=False)))
        r = requests.post(url=url, json=data)
        logger.info("resp : {}".format(r.text))
        try:
            slotvalue = {}
            intentlist = r.json()["intents"][0]["dialog_acts"]
            if len(intentlist) > 0:
                rintent = ""
                for intent in intentlist:
                    rintent = rintent + "," + intent["intent_name"]
                reintents = rintent.strip(",")
            else:
                reintents = "空意图"
            if len(intentlist) > 0:
                for intent in intentlist:
                    # intentslot = {}
                    slotvalue[intent["intent_name"]] = intent["slots"]
            if reintents == "空意图":
                slotvalue = "无槽位"
            return reintents, slotvalue
        except Exception as e:
            print("请求失败：", e)
            return "请求失败:statuscode,{},response {}".format(r.status_code, r.text)

    @staticmethod
    def nlufusionquery(pid, query, env, newnlu, intents=None, proinentlist=None, glbintentlist=None,
                       simintentlist=None):
        logger.info(newnlu)
        nluconfig = newnlu.lower()
        if nluconfig == "true":
            return NluTest.nlufusionaics(query, env, proinentlist, pid, glbintentlist, simintentlist)
        elif nluconfig == "ccdmnlu":
            return NluTest.ccdmnlu(pid, query, env)
        else:
            return NluTest.nlufusioninent(pid, query, env, intents)

    def collectcase(self, wb, sheetname):
        print('获取 sheetname is {}'.format(sheetname))
        sheet1 = wb[sheetname]
        casesexcel = sheet1[self.excelclown['query']]
        pid = str(sheet1[self.config["productid"]].value).strip()
        newnlu = str(sheet1[self.config["nlutype"]].value).strip()
        cases = []
        case = {}
        for idx, val in enumerate(casesexcel[self.excelclown["caserow"]:]):
            if val.value is None:
                pass
            else:
                # print("value is {}".format(val.value))
                case["query"] = str(val.value).strip()
                try:
                    case["expectintents"] = sheet1[self.excelclown["expectintents"] + str(val.row)].value.strip()
                except Exception as e:
                    print("期望意图为空: ", e)
                    case["expectintents"] = "GLB:空意图"
                case["sheetname"] = sheetname
                case["pid"] = pid
                case["row"] = val.row
                case["queryintent"] = sheet1[self.excelclown["queryintet"] + str(val.row)].value
                if case["queryintent"] is not None:
                    case["queryintent"] = case["queryintent"].strip("\n")
                case["filename"] = self.file
                case["expectslots"] = sheet1[self.excelclown["expectslots"] + str(val.row)].value
                case["newnlu"] = newnlu.strip()
                cases.append(deepcopy(case))
                case.clear()
        return cases

    def main(self):
        if len(self.sheetnames) == 0:
            self.sheetnames = self.wb.sheetnames
        cases = []
        for sheetname in self.sheetnames:
            cases.extend(self.collectcase(self.wb, sheetname))
        return cases

    def reset(self):
        """
        重置文本，颜色，移动上次准确率，上次结果
        """
        # 若不给定 表单名则获取全部sheet表进行运行
        if len(self.sheetnames) == 0:
            self.sheetnames = self.wb.sheetnames
        for sheetname in self.sheetnames:
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["query"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["resultintents"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["expectintents"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["comparedresult"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["reexpectslots"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["compareintent"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["expectslots"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["compareslot"], self.excelclown["caserow"])
            operatexcel.resetcolour(self.wb, sheetname, self.excelclown["queryintet"], self.excelclown["caserow"])
            operatexcel.resettext(self.wb, sheetname, self.excelclown["resultintents"], self.excelclown["caserow"])
            operatexcel.resettext(self.wb, sheetname, self.excelclown["comparedresult"], self.excelclown["caserow"])
            operatexcel.resettext(self.wb, sheetname, self.excelclown["reexpectslots"], self.excelclown["caserow"])
            operatexcel.resettext(self.wb, sheetname, self.excelclown["compareintent"], self.excelclown["caserow"])
            operatexcel.resettext(self.wb, sheetname, self.excelclown["compareslot"], self.excelclown["caserow"])
            # self.move_reintent()
            # self.move_precision_rate()

    def casecompare(self):
        """
        根据结果，对行进行  标红操作。并算出准确率
        """
        sheetname_wrong = set()
        if len(self.sheetnames) == 0:
            self.sheetnames = self.wb.sheetnames
        for sheetname in self.sheetnames:
            sheet = self.wb[sheetname]
            compareclown = sheet[self.excelclown["comparedresult"]]
            for idx, val in enumerate(compareclown[self.excelclown["caserow"]:]):
                if val.value is None:
                    pass
                else:
                    if val.value is True:
                        pass
                    else:
                        sheetname_wrong.add(sheetname)
                        val.font = Font(color=RED)
                        sheet[self.excelclown["query"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["resultintents"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["expectintents"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["comparedresult"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["reexpectslots"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["expectslots"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["compareintent"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["compareslot"] + str(val.row)].font = Font(color=RED)
                        sheet[self.excelclown["queryintet"] + str(val.row)].font = Font(color=RED)
            self.get_precision_rate(self.wb, sheetname)
        self.wb.save(self.file)
        with open("test_reuslt", "a", encoding="utf-8") as result_file:
            if len(sheetname_wrong) > 0:
                result_file.write(
                    "运行结束时间：  " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "   执行失败用例表  " + str(
                        sheetname_wrong))
            else:
                result_file.write(
                    "运行结束时间：  " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "   用例全部通过")
            result_file.write("\n")
            print(sheetname_wrong)

    def parsere(self, re):
        # for r in re:
        #     sheet = self.wb[r["sheetname"]]
        #     sheet[self.excelclown["comparedresult"] + str(r["row"])] = r["testre"]
        #     sheet[self.excelclown["resultintents"] + str(r["row"])] = r["reintent"]
        # for r in re:
        print("待解析处理的结果： " + json.dumps(re, ensure_ascii=False))
        sheet = self.wb[re["sheetname"]]
        try:
            sheet[self.excelclown["comparedresult"] + str(re["row"])] = re["testre"]
            sheet[self.excelclown["resultintents"] + str(re["row"])] = str(re["reintent"])
            sheet[self.excelclown["reexpectslots"] + str(re["row"])] = str(re["reslots"])
            sheet[self.excelclown["compareintent"] + str(re["row"])] = re["testre_intent"]
            sheet[self.excelclown["compareslot"] + str(re["row"])] = re["testre_slots"]
        except Exception as e:
            print("结果解析问题： ", e)
            sheet[self.excelclown["comparedresult"] + str(re["row"])] = False

    def get_precision_rate(self, wb, sheetname):
        """
        算出，准确率
        """
        print('获取 sheetname is {}'.format(sheetname))
        sheet1 = wb[sheetname]
        casesexcel = sheet1[self.excelclown['comparedresult']]
        truenum = 0
        totalnum = 0
        for idx, val in enumerate(casesexcel[self.excelclown["caserow"]:]):
            if val.value is None:
                pass
            else:
                if val.value is True:
                    truenum = truenum + 1
                    totalnum = totalnum + 1
                else:
                    totalnum = totalnum + 1
        precision_rate = '{:.2%}'.format(truenum / totalnum)
        # precisionclown = sheet1[self.excelclown['precisionrate']]
        sheet1[self.excelclown['precisionrate'] + str(self.excelclown["caserow"] + 1)].value = precision_rate

    def move_precision_rate(self):
        """
        移动 准确率的值
        """
        if len(self.sheetnames) == 0:
            self.sheetnames = self.wb.sheetnames
        for sheetname in self.sheetnames:
            sheet1 = self.wb[sheetname]
            # casesexcel = sheet1[self.excelclown['precisionrate']]
            for num in reversed(range(1, 7)):
                sheet1[self.excelclown['precisionrate'] + str(self.excelclown["caserow"] + num)].value = sheet1[
                    self.excelclown['precisionrate'] + str(self.excelclown["caserow"] + num - 1)].value

    def save(self):
        # self.move_precision_rate(self.wb,self.sheetnames[0])
        # self.wb.save(self.file)
        # self.move_reintent()
        self.wb.save(self.file)

    def move_reintent(self):
        """
        移动结果
        """
        if self.sheetnames is None:
            self.sheetnames = self.wb.sheetnames
        for sheetname in self.sheetnames:
            sheet1 = self.wb[sheetname]
            # 移动往期结果
            print("移动往期结果")
            for i in reversed(range(ord(self.excelclown["repast_start"]), ord(self.excelclown["repast_end"]))):
                # print(chr(i))
                casesexcel = sheet1[chr(i)]
                for idx, val in enumerate(casesexcel[self.excelclown["caserow"]:]):
                    if val.value is None:
                        pass
                    else:
                        # print("往期移动列，{}, value is {}".format((chr(i + 1) + str(val.row)), val.value))

                        sheet1[chr(i + 1) + str(val.row)].value = val.value
            print("移动往期结果结束")
            print("移动上次结果")
            # 移动上次结果
            reintent = sheet1[self.excelclown["resultintents"]]
            for idx, val in enumerate(reintent[self.excelclown["caserow"]:]):
                if val.value is None:
                    pass
                else:
                    # print("上次移动列，{}, value is {}".format((self.excelclown["repast_start"] + str(val.row)), val.value))
                    sheet1[self.excelclown["repast_start"] + str(val.row)].value = val.value
            print("移动上次结果结束")

    @staticmethod
    def readyaml():
        """
        读取用例配置文件
        """
        file = parent_path2 + "/data/test_nlu/case.yaml"
        with open(file, "r", encoding="utf-8") as conf:
            b = yaml.load(conf.read(), Loader=yaml.FullLoader)

        return b


if __name__ == '__main__':
    file = r"C:\code\platform-regression-python\data\test_nlu\debang.xlsx"
    # r = getexcelsheetconfig(file)
    # print(r)
    runcase(file, threadnum=3)
    # querycase = getcases("sf_quanchangjing_nlu.xlsx")
    # print(querycase)
