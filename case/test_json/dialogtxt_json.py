# encoding:utf-8
import concurrent.futures
import requests
import json
import uuid
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import BLACK
import time
from copy import deepcopy
from base.readpath import getpath
import yaml
import random
from datetime import datetime
class readfileerr(Exception):
    pass


def readfileconf(filename):
    """
    根据文件名返回 & 是否运行 文件属性列表
    :param filename:
    :return:
    """
    file = getpath() + "/data/test_json/test_json.yaml"
    with open(file, "r", encoding="utf-8") as conf:
        b = yaml.load(conf.read(), Loader=yaml.FullLoader)
        filepropertys = []
        for fileproperty in b["filenames"]:
            if fileproperty["file"] == filename and fileproperty["run"] == True:
                fileproperty["file"] = getpath() + "/data/test_json/" + fileproperty["file"]
                print(fileproperty)
                filepropertys.append(fileproperty)
        # if len(filepropertys) > 0:
        #         #     return filepropertys
        #         # else:
        #         #     raise readfileerr("{} 文件不存在".format(filename))
    return filepropertys

def readyaml():
    """
    :return: 获取case 文件配置
    """
    file = getpath() + "/data/test_json/test_json.yaml"
    with open(file, "r", encoding="utf-8") as conf:
        b = yaml.load(conf.read(), Loader=yaml.FullLoader)
        for fileproperty in b["filenames"]:
            fileproperty["file"] = getpath() + "/data/test_json/" + fileproperty["file"]
    return b


class exceloperate():
    caserow = 5  # 用例开始的行数
    # excle 映射
    excelcolumnsmap = {
        # 应答语
        "titleclown": "F",
        # topiccode
        "topiccode": "J",
        # 对话对比结果
        "actioncompare": "I",
        # topic 对比结果
        "topiccompare": "L",
        # 用例综合对比结果
        "casecompare": "H",
        # 步骤query
        "cases": "E",
        # topiccode 预期
        "topiccodeexpect": "K",
        # 话术预期
        "answerexpcet": "G",
        # dataclean-message 参数
        "params": "D",
        "intent": "M"  # 意图列
    }

    def __init__(self, env, phone, excelpath, sheetnames,randphone=False):
        # self.url = self.dmenv[env]
        self.phone = phone
        self.excelpath = excelpath
        self.sheetnames = sheetnames
        # self.caserow=5  #用例开始的行
        self.casenum = 0
        self.env = env
        self.randphone=randphone

    def compareclown(self, wb, sheetname):
        # web=load_workbook(excelpath)
        sheet1 = wb[sheetname]
        # 对话结果处理
        actionclown = sheet1[self.excelcolumnsmap['answerexpcet']]
        for idx, val in enumerate(list(actionclown)[self.caserow:]):
            # print(idx,val)
            if val.value is None:
                pass
            else:
                diaraw = self.excelcolumnsmap['titleclown'] + str(val.row)
                # print(sheet1[diaraw].value)
                actioncompare = self.excelcolumnsmap['actioncompare'] + str(val.row)
                if sheet1[diaraw].value == val.value:
                    sheet1[actioncompare] = 'pass'
                else:
                    sheet1[actioncompare] = 'fail'
                    sheet1[actioncompare].font = Font(color=RED)
        # topiccode 对比处理
        topiccodeexpect = sheet1[self.excelcolumnsmap['topiccodeexpect']]
        for idx, val in enumerate(list(topiccodeexpect)[self.caserow:]):
            # print(idx,val)
            if val.value is None:
                pass
            else:
                topiccode = self.excelcolumnsmap['topiccode'] + str(val.row)
                # print(sheet1[diaraw].value)
                topiccompare = self.excelcolumnsmap['topiccompare'] + str(val.row)
                if str(sheet1[topiccode].value) == str(val.value):
                    sheet1[topiccompare] = 'pass'
                else:
                    sheet1[topiccompare] = 'fail'
                    sheet1[topiccompare].font = Font(color=RED)

        # 汇总对比结果
        dialoguereclown = sheet1[self.excelcolumnsmap['actioncompare']]
        topicreclown = sheet1[self.excelcolumnsmap['topiccompare']]
        rowlist = set()
        for i in list(dialoguereclown)[self.caserow:]:
            if i.value is None:
                pass
            else:
                rowlist.add(i.row)

        for i in list(topicreclown)[self.caserow:]:
            if i.value is None:
                pass
            else:
                rowlist.add(i.row)
        # print(rowlist)
        for row in rowlist:
            dialoguerevalue = sheet1[self.excelcolumnsmap['actioncompare'] + str(row)]
            topicrevalue = sheet1[self.excelcolumnsmap['topiccompare'] + str(row)]
            # print(row,'  ',dialoguerevalue.value,'  ',topicrevalue.value)
            if dialoguerevalue.value == 'fail' or topicrevalue.value == 'fail':
                sheet1[self.excelcolumnsmap['casecompare'] + str(row)] = 'fail'
                sheet1[self.excelcolumnsmap['casecompare'] + str(row)].font = Font(color=RED)
                sheet1[self.excelcolumnsmap['cases'] + str(row)].font = Font(color=RED)
                sheet1[self.excelcolumnsmap["titleclown"] + str(row)].font = Font(color=RED)
                sheet1[self.excelcolumnsmap["answerexpcet"] + str(row)].font = Font(color=RED)
            else:
                sheet1[self.excelcolumnsmap['casecompare'] + str(row)] = 'pass'
        # web.save(excelpath)

    def resetfont(self, wb, sheetname):
        """
        重置用例内的颜色，与结果
        :param wb:
        :param sheetname:
        :return:
        """
        sheet1 = wb[sheetname]
        # 重置case 颜色
        lastcasecompare = sheet1[self.excelcolumnsmap['cases']]
        for idx, val in enumerate(lastcasecompare[self.caserow:]):
            val.font = Font(color=BLACK)
        # 重置 titleclown 颜色
        lasttitleclown = sheet1[self.excelcolumnsmap['titleclown']]
        for idx, val in enumerate(lasttitleclown[self.caserow:]):
            val.font = Font(color=BLACK)
        # 重置answerexpcet 颜色

        lastanswerexpcet = sheet1[self.excelcolumnsmap['answerexpcet']]
        for idx, val in enumerate(lastanswerexpcet[self.caserow:]):
            val.font = Font(color=BLACK)
        # 重置应答语实际结果
        lasttitleclown = sheet1[self.excelcolumnsmap['titleclown']]
        for idx, val in enumerate(lasttitleclown[self.caserow:]):

            if val.value is None:
                pass
            else:
                val.value = ''

        # 重置casecompare 对比结果
        lastcasecompare = sheet1[self.excelcolumnsmap['casecompare']]
        for idx, val in enumerate(lastcasecompare[self.caserow:]):
            val.font = Font(color=BLACK)
            if val.value is None:
                pass
            else:
                val.value = ''

        # 重置topiccodde 结果
        lastcasecompare = sheet1[self.excelcolumnsmap['topiccode']]
        for idx, val in enumerate(lastcasecompare[self.caserow:]):
            val.font = Font(color=BLACK)
            if val.value is None:
                pass
            else:
                val.value = ''

            # 重置topiccompare 对比结果
        lastcasecompare = sheet1[self.excelcolumnsmap['topiccompare']]
        for idx, val in enumerate(lastcasecompare[self.caserow:]):
            val.font = Font(color=BLACK)
            if val.value is None:
                pass
            else:
                val.value = ''

        # 重置 actioncompare对比结果
        lastcasecompare = sheet1[self.excelcolumnsmap['actioncompare']]
        for idx, val in enumerate(lastcasecompare[self.caserow:]):
            val.font = Font(color=BLACK)
            if val.value is None:
                pass
            else:
                val.value = ''
        # 清空 intents 意图列
        lastintents = sheet1[self.excelcolumnsmap['intent']]
        for idx, val in enumerate(lastintents[self.caserow:]):
            # val.font = Font(color=BLACK)
            if val.value is None:
                pass
            else:
                val.value = ''

    def collectcaseba(self, wb, sheetname, phone):
        print('获取 sheetname is {}'.format(sheetname))
        sheet1 = wb[sheetname]
        casesexcel = sheet1[self.excelcolumnsmap['cases']]
        pid = str(sheet1['D2'].value).strip()
        outbound_on_off = str(sheet1['B3'].value).strip()
        outboundparam = str(sheet1['D3'].value)
        # print(cases)
        casedialogue = []
        cases = []
        case = {}
        phonefields = {}
        rephone = phone
        param = {}
        for idx, val in enumerate(casesexcel[self.caserow:]):
            if val.value is None:
                pass
            else:
                case.clear()
                if val.value == False:
                    self.casenum = self.casenum + 1
                    cases.append(deepcopy(casedialogue))
                    casedialogue.clear()
                    param.clear()
                    if self.randphone == True:
                        rephone = "185"+str(str(int(time.time() * 10))[-5:] + str(int(random.random() * 10000)))
                    else:
                        rephone = self.phone
                    phone = rephone
                    paramsclown = self.excelcolumnsmap['params'] + str(val.row)
                    params = sheet1[paramsclown].value
                    paramdict = self.parseparamsba(params)
                    if paramdict is None:
                        pass
                    else:
                        if paramdict.get("phone") is not None:
                            phone = paramdict.get("phone")
                            param['callNum'] = phone
                        if paramdict.get("phonefield") is not None:
                            if phonefields.get(paramdict.get("phonefield")) is None:
                                phone="185"+str(str(int(time.time() * 10))[-5:] + str(int(random.random() * 10000)))
                                phonefields[paramdict.get("phonefield")]=phone
                            else:
                                phone=phonefields[paramdict.get("phonefield")]
                    uustr = str(uuid.uuid1()).replace('-', '')
                    senderId = phone + '&' + uustr
                    param['callNum'] = phone
                    sessionId = uustr
                case['sheetname'] = sheetname
                case['row'] = val.row
                case['senderId'] = senderId
                case['sessionId'] = sessionId
                case['pid'] = pid
                case["outbound"] = outbound_on_off
                case["ourbound_message"] = outboundparam
                case["filename"] = self.excelpath
                case["env"] = self.env
                case["answerexpcet"] = sheet1[self.excelcolumnsmap["answerexpcet"] + str(val.row)].value
                case["topiccodeexpect"] = sheet1[self.excelcolumnsmap["topiccodeexpect"] + str(val.row)].value
                paramsclown = self.excelcolumnsmap['params'] + str(val.row)
                params = sheet1[paramsclown].value
                paramdict = self.parseparamsba(params)

                if paramdict is None:
                    param['command'] = 0
                else:
                    if paramdict.get("command") is not None:
                        command = paramdict.get("command")
                        param['command'] = command
                    else:
                        param['command'] = 0

                param['query'] = (str(val.value))
                # 开始语句确保初始化command 为 1 newcall
                if param['query'] == 'False':
                    param['command'] = 1

                if param['query'] == 'asr-silence':
                    param['command'] = 4

                if param['query'] == 'hangup':
                    param['command'] = 2
                case['param'] = deepcopy(param)
                casedialogue.append(deepcopy(case))

        cases.append(deepcopy(casedialogue))
        cases.pop(0)
        # print(cases)
        phonefields.clear()
        return cases

    def collectquery(self, wb, sheetname):
        """
        获取所有待查询意图语料
        :param wb:
        :param sheetname:
        :return: querys
        """
        # return querys
        # print('获取的 ')
        sheet1 = wb[sheetname]
        queryexcel = sheet1[self.excelcolumnsmap['cases']]
        sw = str(sheet1['B4'].value)  # 是否查询意图
        # print(bool(sw))
        # print(type(sw))
        method = str(sheet1['D4'].value).strip()
        if sw == 'True':
            # print(querycases)
            querycasedialogue = []
            querycases = []
            case = {}
            for idx, val in enumerate(queryexcel[self.caserow:]):
                if val.value is None:
                    pass
                else:
                    case.clear()
                    # print("value is ",val.value)
                    if val.value == False:
                        querycases.append(deepcopy(querycasedialogue))
                        querycasedialogue.clear()
                    else:
                        case['sheetname'] = sheetname
                        case['row'] = val.row
                        case["method"] = method
                        case["query"] = val.value.strip()
                        case["filename"] = self.excelpath
                        case["env"] = self.env
                        querycasedialogue.append(deepcopy(case))

            querycases.append(deepcopy(querycasedialogue))
            querycases.pop(0)
            # print(querycases)
            return querycases
        else:
            return []

    def parseparamsba(self, params):
        """
        解析params单元格内自定义参数: param1:value,param:value,解析 dataclean-message 类型的参数
        :param params:
        :return:
        """
        param = {}
        try:
            for item in params.strip().split(','):
                param[item.split(':')[0]] = item.split(':')[1]
            if param.get('command') is None:
                param['command'] = 0
        except:
            params is None
        return param

    def handledmresult(self, wb, cases):
        # sheet1 = wb[sheetname]
        for casegroup in cases:

            for case in casegroup:
                sheet1 = wb[case['sheetname']]
                # print('handleresult case.yaml is{}'.format(case))
                row = str(case['row'])
                # result = case['result']

                # 应答语
                titleclown = self.excelcolumnsmap['titleclown'] + row
                # topiccode
                topiccodeclown = self.excelcolumnsmap["topiccode"] + row
                sheet1[titleclown] = case.get("titleclown")
                sheet1[topiccodeclown] = case.get("topiccode")
                # resultdict = json.loads(result)
                # try:
                #     sheet1[titleclown] = resultdict['voiceText'].strip()
                #     # topiccode
                #     topiccde = resultdict["topicCode"]
                #     sheet1[topiccodeclown] = topiccde
                # except:
                #     pass

    def handleintentresult(self, wb, querys):
        """
        处理意图的结果，把意图请求的list 中intent 填入 excel
        :param wb:
        :param querys:
        :return:
        """
        for queryitme in querys:
            for query in queryitme:
                sheet1 = wb[query['sheetname']]
                # print('handleresult case.yaml is{}'.format(query))
                row = str(query['row'])
                result = query['intent']

                # 意图列表
                intentclown = self.excelcolumnsmap['intent'] + row
                sheet1[intentclown] = result

    def main(self):
        start_time = time.perf_counter()
        wb = load_workbook(self.excelpath)
        print(wb.sheetnames)
        # 若不给定 表单名则获取全部sheet表进行运行
        if len(self.sheetnames) == 0:
            self.sheetnames = wb.sheetnames
        # 重置字体颜色
        for sheetname in self.sheetnames:
            self.resetfont(wb=wb, sheetname=sheetname)

        cases = []
        querys = []
        for sheetname in self.sheetnames:
            cases.extend(self.collectcaseba(wb=wb, sheetname=sheetname, phone=self.phone))
            querys.extend(self.collectquery(wb=wb, sheetname=sheetname))
        print('cases is {}'.format(cases))
        print("querys is {}".format(querys))
        self.postdm_all(cases)
        self.handledmresult(wb, cases)
        if len(querys) > 0:
            self.postintent_all(querys)
            self.handleintentresult(wb, querys)
        # 循环标记结果
        for sheetname in self.sheetnames:
            self.compareclown(wb, sheetname)

        wb.save(self.excelpath)

        # pass
        endtime = time.perf_counter()
        print('{} time cost {}'.format(self.excelpath, (endtime - start_time)))

    def getcase(self):
        """
        根据 文件名，表名获取对应的用例
        :return:
        """
        wb = load_workbook(self.excelpath)
        print(wb.sheetnames)
        # 若不给定 表单名则获取全部sheet表进行运行
        if len(self.sheetnames) == 0:
            self.sheetnames = wb.sheetnames
        cases = []
        # querys = []
        for sheetname in self.sheetnames:
            cases.extend(self.collectcaseba(wb=wb, sheetname=sheetname, phone=self.phone))
            # querys.extend(self.collectquery(wb=wb, sheetname=sheetname))
        return cases

    def getquerys(self):
        """
               根据 文件名，表名获取对应的查询语料
               :return:
               """
        wb = load_workbook(self.excelpath)
        print(wb.sheetnames)
        # 若不给定 表单名则获取全部sheet表进行运行
        if len(self.sheetnames) == 0:
            self.sheetnames = wb.sheetnames
        querys = []
        for sheetname in self.sheetnames:
            querys.extend(self.collectquery(wb=wb, sheetname=sheetname))
        return querys

    def before_execute(self):
        """重置excel格式"""
        wb = load_workbook(self.excelpath)
        for sheetname in self.sheetnames:
            self.resetfont(wb=wb, sheetname=sheetname)
        wb.save(self.excelpath)

    def after_execute(self, cases, querys):
        """
        结果回填，与结果对比
        :param cases:
        :param querys:
        :return:
        """
        wb = load_workbook(self.excelpath)
        self.handledmresult(wb, cases)
        if len(querys) > 0:
            self.handleintentresult(wb, querys)
        # 循环标记结果
        for sheetname in self.sheetnames:
            self.compareclown(wb, sheetname)
        # wb.save(self.excelpath)
        wb.save(self.excelpath)


class postservice():
    dm_runtime_url = {
        "test": "http://dm-runtime-test.talkinggenie.com/callcenter/nlu",
        "alpha": "http://dm-runtime-alpha.talkinggenie.com/callcenter/nlu",
        "beta": "http://dm-runtime-beta.talkinggenie.com/callcenter/nlu",
    }
    baenv = {
        'test': 'http://nlu-test.talkinggenie.com/dataclean/message/yto/messages/v2',
        'alpha': 'http://nlu-alpha.talkinggenie.com/dataclean/message/yto/messages/v2',
        "beta": "http://debugcheck-beta.talkinggenie.com/dataclean/message/yto/messages/v2"
    }
    aihiveboxurl = {
        "test": "http://ba-bnlu-test.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7",
        "alpha": "http://ba-bnlu-alpha.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7",
        "beta": "http://ba-bnlu-beta.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7"

    }
    dm_runtime_url = {
        "test": "http://dm-runtime-test.talkinggenie.com/callcenter/nlu",
        "alpha": "http://dm-runtime-alpha.talkinggenie.com/callcenter/nlu",
        "beta": "http://dm-runtime-beta.talkinggenie.com/callcenter/nlu",
    }

    def postba(self, case):
        """
        对dataclean_message 请求
        :param case:
        :return:
        """
        url = self.baenv[case["env"]]
        # print('url is {}'.format(url))
        # print('case.yaml is {}'.format(case))
        body = {
            "content": case['param']['query'],
            "senderId": case['senderId'],
            "externalSessionId": "",
            "productId": case['pid'],
            "recordId": str(uuid.uuid1()).replace('-', ''),
            "expressNumber": "",
            "command": case['param']['command'],
            "type": "1",
            "isPlaying": "false",
            "playing": {
                "content": "",
                "type": ""
            },
            "recordingFile": "",
            "startTime": datetime.now().strftime('%Y-%m-%d-%H-%M-%S'),
            "stopTime": datetime.now().strftime('%Y-%m-%d-%H-%M-%S'),
            "asrConfidence": "1.0",
            "callNum": case['param']['callNum'],
            "sessionId": case['sessionId'],
            "outboundStartTime": "",
            "outboundEndTime": "",
            "outboundStatus": "",
            "destNumber": "",
            "hangupReason": "other"
        }
        # print('body is {}'.format(body))
        result = requests.post(url=url, json=body)
        # print("senderId %s", case['senderId'])
        # print("request %s", body)
        # print("status_code %d status_text %s" % (result.status_code, result.text))
        return result

    def postintent(self, querys):
        print("query is ", querys)
        for querycasegroup in querys:
            for query in querycasegroup:
                if query["method"] == "aichov":
                    url = self.aihiveboxurl[query["env"]]
                    r = self.query_ba_aihivebox(query["query"], url)
                    query["intent"] = r

    def query_ba_aihivebox(self, query, url):
        """
        query 意图查询语料，
        :param query:
        :return: str,意图名字 "name1,name2"
        """
        dataraw = {"params": {"request": {"refText": query, "coreType": "cn.dlg.ita", "res": "aihivebox",
                                          "env": "use_slot_index=1"}}}
        data = "params" + "=" + json.dumps(dataraw["params"], ensure_ascii=False)
        print(data)
        try:
            r = requests.post(url=url, data=data.encode("utf-8"))
            if r.status_code == 200:
                param = json.loads(r.text)["result"]["semantics"]["request"]["param"]
                rstr = ''
                for k, v in param.items():
                    rstr = rstr + str(k) + ":" + str(v) + ","
                    return rstr.strip(",")
            else:
                return "意图参数问题，返回非200"
            # print(r.text)
        except:
            return "意图请求错误"

    def postoutbound(self, case):
        """
        发起外呼请求
        :param case:
        :return:
        """
        data = {
            "senderId": case["senderId"],
            "sessionId": case["sessionId"],
            "asrConfidence": 1,
            "productId": case["pid"],
            "query": "",
            "message": json.loads(case["ourbound_message"])
        }
        data["message"]["appId"] = int(data["message"]["appId"])
        # print("outbounder request is   ", json.dumps(data, ensure_ascii=False))
        outboundr = requests.post(url=self.dm_runtime_url[case["env"]], json=data)
        # print("outbounder status_code is ", outboundr.status_code)
        # print("ourbounder response is ", outboundr.text)
        return outboundr

    def postall(self, casegroup):
        """
        进行：内呼，外呼
        :param casegroup:
        :return:
        """
        # print('casegroup is {}'.format(casegroup))

        for case in casegroup:
            if case["outbound"] == "True" and case["param"]["query"].lower() == 'false':
                outboundr = self.postoutbound(case)
                result = self.postba(case)
                case["titleclown"] = resultdict['voiceText'].strip()
                case["topiccode"] = resultdict["topicCode"]
                # try:
                #     if json.loads(outboundr.text)["best"]["reply"]["topicCode"] == "1007":
                #         result = self.postba(case)
                #         # case['result'] = result.text
                #         # print('response is {}'.format(result.text))
                #         resultdict = json.loads(result.text)
                #         case["titleclown"] = resultdict['voiceText'].strip()
                #         # topiccode
                #         case["topiccode"] = resultdict["topicCode"]
                #
                #     else:
                #         # result = self.postba(case)
                #         case['titleclown'] = "发起外呼失败"
                #         case["topiccode"] = ""
                #         # print('response is {}'.format(result.text))
                # except:
                #     pass
            else:
                result = self.postba(case)
                resultdict = json.loads(result.text)
                case["titleclown"] = resultdict['voiceText'].strip()
                # topiccode
                case["topiccode"] = resultdict["topicCode"]
                # print('response is {}'.format(result.text))


if __name__ == '__main__':
    pass
    from case.test_json.test_json import getcase
    a=getcase("debangba_alpha.xlsx")
    print("**************************")
    print(a)