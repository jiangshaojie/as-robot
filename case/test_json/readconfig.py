# -*- coding: UTF-8 -*-
from openpyxl import load_workbook
import datetime


class ReadConfig:
    excel_bot_config = {
        "env": "F3",
        "run": "F4",
        "phone": "H3",
        "time": "H4"
    }

    def __init__(self, filename):
        self.filename = filename

    def readbot(self):
        wb = load_workbook(self.filename)
        configs = []
        runconfig = {}
        # runconfig[self.filename]={}
        for sheetname in wb.sheetnames:
            configs.append(self.getbotconfig(wb, sheetname))
        for config in configs:
            if len(config) > 0:
                if runconfig.get(config.get("env")) is None:
                    runconfig[config.get("env")] = []
                    runconfig[config.get("env")].append(config.get("sheetname"))
                else:
                    runconfig[config.get("env")].append(config.get("sheetname"))
        return runconfig

    def getbotconfig(self, wb, sheetname):
        sheet = wb[sheetname]
        env = sheet[self.excel_bot_config.get("env")].value
        if env is not None:
            env = env.strip()
        run = sheet[self.excel_bot_config.get("run")].value
        phone = sheet[self.excel_bot_config.get("phone")].value
        time = sheet[self.excel_bot_config.get("time")].value
        # print(time)
        timekeeper = datetime.datetime.now().strftime("%H:%M:%S")
        if not isinstance(time, type(None)):
            runtime = time.strip().strip("[").strip("]").replace("\"", "").split(",")
            if timekeeper < runtime[0] or timekeeper > runtime[1]:
                return {}
        if run:
            return {"env": env, "phone": phone, "sheetname": sheetname}
        else:
            return {}


if __name__ == '__main__':
    file = "C:/code/platform-regression-python/data/test_json/yunpinlvtest.xlsx"
    a = ReadConfig(file)
    # a.getbotconfig()
    a.readbot()
