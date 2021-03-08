# encoding:utf-8
import concurrent.futures
import requests
import json
import uuid
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import BLACK
import time
from copy import deepcopy
from copy import copy
import random
from datetime import datetime
from datetime import timedelta
from base.log import logger
from case.test_json.acquirekibana import acquirekibana


class dialogtxt():
    caserow = 5  # 用例开始的行数
    # url 访问列表N
    baenv = {
        'test': 'http://nlu-test.talkinggenie.com/dataclean/message/yto/messages/v2',
        'alpha': 'http://nlu-alpha.talkinggenie.com/dataclean/message/yto/messages/v2',
        "beta": "http://debugcheck-beta.talkinggenie.com/dataclean/message/yto/messages/v2",
        "dev": "http://nlu-dev.talkinggenie.com/dataclean/message/yto/messages/v2"
    }
    aihiveboxurl = {
        "test": "http://ba-bnlu-test.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7",
        "alpha": "http://ba-bnlu-alpha.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7",
        "beta": "http://ba-bnlu-beta.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7",
        "dev": "http://ba-bnlu-dev.talkinggenie.com/cn.dlg.ita/aihivebox?__internal__=1&version=0.0.7"

    }
    dm_runtime_url = {
        "test": "http://dm-runtime-test.talkinggenie.com/callcenter/nlu",
        "alpha": "http://dm-runtime-alpha.talkinggenie.com/callcenter/nlu",
        "beta": "http://dm-runtime-beta.talkinggenie.com/callcenter/nlu",
        "dev": "http://dm-runtime-dev.talkinggenie.com/callcenter/nlu",
    }
    # excle 映射
    excelcolumnsmap = {
        # 应答语
        "titleclown": "F",
        # topicode
        "topiccodde": "J",
        # 对话对比结果
        "actioncompare": "I",
        # topic 对比结果
        "topiccompare": "L",
        # 用例综合对比结果
        "casecompare": "H",
        # 步骤query
        "cases": "E",
        # topiccode 预期
        "topiccodeexpect": "K",
        # 话术预期
        "answerexpcet": "G",
        # dataclean-message 参数
        "params": "D",
        # "intent": "P",  # 意图列
        "transferNum": "M",
        "transferNumexpect": "N",
        "transferNumcompare": "O",
        "labels": "P",
        "labelsexpect": "Q",
        "labelscompare": "R",
        "sessionId": "S",
        "servicetype": "T",
        "servicetype_expect": "U",
        "servicetype_compare": "V",
        "topic_transfer": "W",
        "orderNum": "X",
        "orderNum_expcet": "Y",
        "orderNum_compare": "Z"
    }

    def __init__(self, excelpath, env=None, phone=None, sheetnames=None, randphone=False, threadnum=None,
                 check_topic_type=None):

        # self.url = self.dmenv[env]
        self.phone = phone
        self.excelpath = excelpath
        self.sheetnames = sheetnames
        self.env = env
        self.url = self.baenv[env]
        self.aihivebox_url = self.aihiveboxurl[env]
        self.dmruntimeurl = self.dm_runtime_url[env]
        # self.caserow=5  #用例开始的行
        self.casenum = 0
        self.randphone = randphone
        self.threadnum = threadnum
        self.check_topic_type = check_topic_type
        self.sheetname_wrong = set()

    def compareclown(self, wb, sheetname):
        # web=load_workbook(excelpath)
        sheet1 = wb[sheetname]
        # 对话结果处理
        actionclown = sheet1[self.excelcolumnsmap['answerexpcet']]
        for idx, val in enumerate(list(actionclown)[self.caserow:]):
            # print(idx,val)
            if val.value is None:
                pass
            else:
                diaraw = self.excelcolumnsmap['titleclown'] + str(val.row)
                # print(sheet1[diaraw].value)
                actioncompare = self.excelcolumnsmap['actioncompare'] + str(val.row)
                if sheet1[diaraw].value.strip() == val.value.strip():
                    sheet1[actioncompare] = 'pass'
                else:
                    sheet1[actioncompare] = 'fail'
                    sheet1[actioncompare].font = Font(color=RED)
        # transernum 对比处理
        transernumexpect = sheet1[self.excelcolumnsmap['transferNumexpect']]
        for idx, val in enumerate(list(transernumexpect)[self.caserow:]):
            # print(idx,val)
            if val.value is None:
                pass
            else:
                transernum = self.excelcolumnsmap['transferNum'] + str(val.row)
                # print(sheet1[diaraw].value)
                transernumcompare = self.excelcolumnsmap['transferNumcompare'] + str(val.row)
                if str(sheet1[transernum].value) == str(val.value):
                    sheet1[transernumcompare] = 'pass'
                else:
                    sheet1[transernumcompare] = 'fail'
                    sheet1[transernumcompare].font = Font(color=RED)
        # topiccode 对比处理
        topiccodeexpect = sheet1[self.excelcolumnsmap['topiccodeexpect']]
        for idx, val in enumerate(list(topiccodeexpect)[self.caserow:]):
            # print(idx,val)
            if val.value is None:
                pass
            else:
                topiccode = self.excelcolumnsmap['topiccodde'] + str(val.row)
                # print(sheet1[diaraw].value)
                topiccompare = self.excelcolumnsmap['topiccompare'] + str(val.row)
                if str(sheet1[topiccode].value) == str(val.value):
                    sheet1[topiccompare] = 'pass'
                else:
                    sheet1[topiccompare] = 'fail'
                    sheet1[topiccompare].font = Font(color=RED)
        # 标签对比结果
        labelsexpect = sheet1[self.excelcolumnsmap['labelsexpect']]
        for idx, val in enumerate(list(labelsexpect)[self.caserow:]):
            # print(idx,val)
            if val.value is None:
                pass
            else:
                labels = self.excelcolumnsmap['labels'] + str(val.row)
                # print(sheet1[diaraw].value)
                labelscompare = self.excelcolumnsmap['labelscompare'] + str(val.row)
                if sheet1[labels].value == None:
                    labelsjson = {}
                else:
                    labelsjson = json.loads(sheet1[labels].value)
                labelsexpectjson = json.loads(val.value)
                flag = True
                for k, v in labelsexpectjson.items():
                    try:
                        if labelsjson.get(k).strip() == v.strip():
                            pass
                        else:
                            flag = False
                            break
                    except:
                        flag = False

                if flag:
                    sheet1[labelscompare] = 'pass'
                else:
                    sheet1[labelscompare] = 'fail'
                    sheet1[labelscompare].font = Font(color=RED)

        # servicetype
        servicetype_expect = sheet1[self.excelcolumnsmap['servicetype_expect']]
        for idx, val in enumerate(list(servicetype_expect)[self.caserow:]):
            # print(idx,val)
            if val.value is None:
                pass
            else:
                servicetype = self.excelcolumnsmap['servicetype'] + str(val.row)
                # print(sheet1[diaraw].value)
                servicetype_compare = self.excelcolumnsmap['servicetype_compare'] + str(val.row)
                if str(sheet1[servicetype].value) == str(val.value):
                    sheet1[servicetype_compare] = 'pass'
                else:
                    sheet1[servicetype_compare] = 'fail'
                    sheet1[servicetype_compare].font = Font(color=RED)
        # topic_transernumr
        topiccode_clown = sheet1[self.excelcolumnsmap['topiccodde']]
        for idx, val in enumerate(list(topiccode_clown)[self.caserow:]):
            # print(idx,val)
            if val.value is None or val.value == "":
                pass
            else:
                if str(val.value) == "1002":
                    servicetypeclown = sheet1[self.excelcolumnsmap['servicetype'] + str(val.row)]
                    transernumclown = sheet1[self.excelcolumnsmap['transferNum'] + str(val.row)]
                    topic_transferclown = sheet1[self.excelcolumnsmap["topic_transfer"] + str(val.row)]
                    if self.check_topic_type == "transernum_servicetype":
                        if servicetypeclown.value is None or transernumclown.value is None:
                            topic_transferclown.value = "fail"
                            topic_transferclown.font = Font(color=RED)
                    elif self.check_topic_type == "servicetype":
                        if servicetypeclown.value is None:
                            topic_transferclown.value = "fail"
                            topic_transferclown.font = Font(color=RED)
                    else:
                        if transernumclown.value is None:
                            topic_transferclown.value = "fail"
                            topic_transferclown.font = Font(color=RED)

        # orderNum对比，目前只有顺丰项目用到了
        topiccode_clown = sheet1[self.excelcolumnsmap['orderNum_expcet']]
        for idx, val in enumerate(list(topiccode_clown)[self.caserow:]):
            # print(idx,val)
            if val.value is None or val.value == "":
                pass
            else:
                order_num = self.excelcolumnsmap['orderNum'] + str(val.row)
                # print(sheet1[diaraw].value)
                ordernum_compare = self.excelcolumnsmap['orderNum_compare'] + str(val.row)
                if str(sheet1[order_num].value) == str(val.value):
                    sheet1[ordernum_compare].value = 'pass'
                else:
                    sheet1[ordernum_compare].value = 'fail'
                    sheet1[ordernum_compare].font = Font(color=RED)
        # 汇总对比结果
        dialoguereclown = sheet1[self.excelcolumnsmap['actioncompare']]
        topicreclown = sheet1[self.excelcolumnsmap['topiccompare']]
        transernumreclown = sheet1[self.excelcolumnsmap["transferNumcompare"]]
        labelscompareclown = sheet1[self.excelcolumnsmap["labelscompare"]]
        servicetype_comparecolwn = sheet1[self.excelcolumnsmap['servicetype_compare']]
        topic_transferclown = sheet1[self.excelcolumnsmap["topic_transfer"]]
        ordernum_compreclown = sheet1[self.excelcolumnsmap["orderNum_compare"]]
        rowlist = set()
        clowns = [dialoguereclown, topicreclown, transernumreclown, labelscompareclown, servicetype_comparecolwn,
                  topic_transferclown, ordernum_compreclown]
        for clown in clowns:
            for i in list(clown)[self.caserow:]:
                if i.value is None:
                    pass
                else:
                    rowlist.add(i.row)
        # print(rowlist)
        for row in rowlist:
            dialoguerevalue = sheet1[self.excelcolumnsmap['actioncompare'] + str(row)]
            topicrevalue = sheet1[self.excelcolumnsmap['topiccompare'] + str(row)]
            transernumrevalue = sheet1[self.excelcolumnsmap["transferNumcompare"] + str(row)]
            labelscomparevalue = sheet1[self.excelcolumnsmap["labelscompare"] + str(row)]
            servicetype_comparevalue = sheet1[self.excelcolumnsmap["servicetype_compare"] + str(row)]
            topic_transfer_value = sheet1[self.excelcolumnsmap["topic_transfer"] + str(row)]
            ordernum_comparevlue = sheet1[self.excelcolumnsmap["orderNum_compare"] + str(row)]
            # print(row, '  ', dialoguerevalue.value, '  ', topicrevalue.value, ' ', transernumrevalue.value)
            if dialoguerevalue.value == 'fail' or topicrevalue.value == 'fail' or transernumrevalue.value == 'fail' or labelscomparevalue.value == 'fail' or servicetype_comparevalue.value == 'fail' \
                    or topic_transfer_value.value == "fail" or ordernum_comparevlue.value == "fail":
                self.sheetname_wrong.add(sheetname)
                sheet1[self.excelcolumnsmap['casecompare'] + str(row)].value = 'fail'
                sheet1[self.excelcolumnsmap['casecompare'] + str(row)].font = Font(color=RED)
                sheet1[self.excelcolumnsmap['cases'] + str(row)].font = Font(color=RED)
                sheet1[self.excelcolumnsmap["titleclown"] + str(row)].font = Font(color=RED)
                sheet1[self.excelcolumnsmap["answerexpcet"] + str(row)].font = Font(color=RED)
                # sheet1[self.excelcolumnsmap["answerexpcet"] + str(row)].font = Font(color=RED)
            else:
                sheet1[self.excelcolumnsmap['casecompare'] + str(row)].value = 'pass'
        # web.save(excelpath)

    def resetfont(self, wb, sheetname):
        """
        重置用例内的颜色，与结果
        :param wb:
        :param sheetname:
        :return:
        """
        sheet1 = wb[sheetname]
        # 重置颜色
        clowns = ["cases", "titleclown", "answerexpcet", "topic_transfer", "orderNum_compare"]
        for clown in clowns:
            for idx, val in enumerate(sheet1[self.excelcolumnsmap[clown]][self.caserow:]):
                val.font = Font(color=BLACK)

        # 重置结果
        clowns = ["titleclown", "casecompare", "topiccodde", "transferNum", "labels", "topiccompare", "actioncompare",
                  "transferNumcompare", "labelscompare", "sessionId", "servicetype", "servicetype_compare",
                  "topic_transfer", "orderNum", "orderNum_compare"]
        for clown in clowns:
            for idx, val in enumerate(sheet1[self.excelcolumnsmap[clown]][self.caserow:]):
                val.font = Font(color=BLACK)
                if val.value is None:
                    pass
                else:
                    val.value = ''

    @staticmethod
    def getPhoneNum():
        pre = "186"
        phoneList = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        for i in range(8):
            pre += (str)(phoneList[random.randrange(0, 10, 1)])
        print(pre)
        return pre

    def gettailnum(self):
        number = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        n = 0
        tailnum = ""
        while (n < 4):
            n += 1
            tailnum += str(random.choice(number))
        return tailnum

    def collectcaseba(self, wb, sheetname, phonenumber):
        logger.info('获取 用例的sheetname： {}'.format(sheetname))
        sheet1 = wb[sheetname]
        casesexcel = sheet1[self.excelcolumnsmap['cases']]
        pid = str(sheet1['D2'].value).strip()
        conf = str(sheet1['B4'].value).strip()
        config_params = None
        if conf.lower() == "true":
            config_params = json.loads(str(sheet1['D4'].value))  # 默认ba配置参数
            print(config_params)
        checklable_on_off = str(sheet1['F2'].value)  # 是否进行标签验证配置
        outbound_on_off = str(sheet1['B3'].value).strip()
        if outbound_on_off == "True":
            outboundparam = json.loads(str(sheet1['D3'].value))
        else:
            outboundparam = str(sheet1['D3'].value)
        outboundconfig_on_off = str(sheet1['F3'].value).strip()
        # print(cases)
        excel_phone_config = sheet1['H3'].value
        if excel_phone_config is not None:
            phonenumber = str(excel_phone_config).strip()
        casedialogue = []
        cases = []
        case = {}
        phonefields = {}
        param = {}
        kwargs = {}
        for idx, val in enumerate(casesexcel[self.caserow:]):
            if val.value is None:
                pass
            else:
                case.clear()
                paramsclown = self.excelcolumnsmap['params'] + str(val.row)
                params = sheet1[paramsclown].value
                paramdict = self.parseparamsba(params)
                if str(val.value).lower() == "false":
                    kwargs.clear()
                    if config_params is not None:
                        print("sw is", config_params)
                        kwargs.update(config_params)
                    self.casenum = self.casenum + 1
                    cases.append(deepcopy(casedialogue))
                    casedialogue.clear()
                    param.clear()
                    if self.randphone == True:
                        rephone = "185" + str(str(int(time.time() * 10))[-4:] + self.gettailnum())
                    else:
                        rephone = phonenumber
                    phone = rephone

                    if paramdict is None:
                        pass
                    else:
                        if paramdict.get("phone") is not None:
                            phone = paramdict.pop("phone")
                            param['callNum'] = phone
                        if paramdict.get("phonefield") is not None:
                            if phonefields.get(paramdict.get("phonefield")) is None:
                                phone = "185" + str(str(int(time.time() * 10))[-4:] + self.gettailnum())
                                phonefields[paramdict.pop("phonefield")] = phone
                            else:
                                phone = phonefields[paramdict.pop("phonefield")]
                        if paramdict.get("cityCode") is not None:
                            citycode = paramdict.pop("cityCode")
                            kwargs["cityCode"] = citycode
                        # else:
                        #     kwargs["citycode"] = ""
                        if paramdict.get("callduration") is not None:
                            case['callduration'] = paramdict.pop("callduration")
                        if paramdict.get("timeago") is not None:
                            case["timeago"] = paramdict.pop("timeago")
                    uustr = phone + str(uuid.uuid1()).replace('-', '')
                    senderId = phone + '&' + uustr
                    param['callNum'] = phone
                    sessionId = uustr
                case["checklable_on_off"] = checklable_on_off
                case['sheetname'] = sheetname
                case['row'] = val.row
                case['senderId'] = senderId
                case['sessionId'] = sessionId
                case['pid'] = pid
                case["outbound"] = outbound_on_off
                case["ourbound_message"] = outboundparam
                if outboundconfig_on_off == "True":
                    outboundconfig_param = json.loads(str(sheet1['H3'].value))
                else:
                    outboundconfig_param = {"taskId": "", "phoneId": ""}
                case["outboundconfig_param"] = outboundconfig_param
                # case["outboundconfig_on_off"]=outbound_on_off

                if pid == "914009574":  # 顺丰主动服务标识
                    # case["kwargs"] = {}
                    flag = paramdict.get("flag")
                    if flag is None:
                        # case["kwargs"]["flag"] = 0
                        kwargs["flag"] = 0
                    else:
                        # case["kwargs"]["flag"] = int(flag)
                        kwargs["flag"] = int(flag)

                if paramdict.get("command") is not None:
                    command = paramdict.pop("command")
                    param['command'] = command
                else:
                    param['command'] = "0"

                param['query'] = (str(val.value))
                # 开始语句确保初始化command 为 1 newcall
                if param['query'].lower() == 'false':
                    param['command'] = "1"
                    param["query"] = "false"

                if param['query'] == 'asr-silence':
                    param['command'] = "4"

                if param['query'] == 'hangup':
                    param["query"] = "0"
                    param['command'] = "2"
                if param["query"].lower() == "signal=hangup" and checklable_on_off == "True":
                    case['lablesparam'] = json.loads(str(sheet1['H2'].value))
                case["kwargs"] = copy(kwargs)
                case['param'] = deepcopy(param)
                casedialogue.append(deepcopy(case))

        cases.append(deepcopy(casedialogue))
        cases.pop(0)
        logger.info("获取的用例：{}".format(cases))
        phonefields.clear()
        return cases

    def collectquery(self, wb, sheetname):
        """
        获取所有待查询意图语料
        :param wb:
        :param sheetname:
        :return: querys
        """
        # return querys
        # print('获取的 ')
        sheet1 = wb[sheetname]
        queryexcel = sheet1[self.excelcolumnsmap['cases']]
        # sw = str(sheet1['B4'].value)  # 是否查询意图  暂时注销掉不用意图配置，原配置给dataclean-message默认请求参数用
        sw = False
        # print(bool(sw))
        # print(type(sw))
        method = str(sheet1['D4'].value).strip()
        if sw == 'True':
            # print(querycases)
            querycasedialogue = []
            querycases = []
            case = {}
            for idx, val in enumerate(queryexcel[self.caserow:]):
                if val.value is None:
                    pass
                else:
                    case.clear()
                    # print("value is ",val.value)
                    if val.value == False:
                        querycases.append(deepcopy(querycasedialogue))
                        querycasedialogue.clear()
                    else:
                        case['sheetname'] = sheetname
                        case['row'] = val.row
                        case["method"] = method
                        case["query"] = val.value.strip()
                        querycasedialogue.append(deepcopy(case))

            querycases.append(deepcopy(querycasedialogue))
            querycases.pop(0)
            # print(querycases)
            return querycases
        else:
            return []

    def parseparamsba(self, params):
        """
        解析params单元格内自定义参数: param1:value,param:value,解析 dataclean-message 类型的参数
        :param params:
        :return:
        """
        param = {}
        try:
            for item in params.strip().split(','):
                param[item.split(':')[0]] = item.split(':')[1]
            if param.get('command') is None:
                param['command'] = "0"
        except:
            params is None
        return param

    def postoutbound(self, case):
        """
        发起外呼请求
        :param case:
        :return:
        """
        data = {
            "senderId": case["senderId"],
            "sessionId": case["sessionId"],
            "asrConfidence": 1,
            "productId": case["pid"],
            "query": "signal=newCall",
            "message": case["ourbound_message"]
        }
        data["message"]["appId"] = int(data["message"]["appId"])
        print("outbounder request is   ", json.dumps(data, ensure_ascii=False))
        outboundr = requests.post(url=self.dmruntimeurl, json=data)
        print("outbounder status_code is ", outboundr.status_code)
        print("ourbounder response is ", outboundr.text)
        return outboundr

    def postba(self, case, start_time=None, stop_time=None):
        """
        对dataclean_message 请求
        :param stop_time:
        :param start_time:
        :param case:
        :return:
        """
        if start_time is None:
            starttime = datetime.now()
        else:
            starttime = start_time
        if stop_time is None:
            stoptime = datetime.now()
        else:
            stoptime = stop_time
        url = self.url
        logger.info('url is {}'.format(url))
        logger.info('case.yaml is {}'.format(case))
        print("starttime and stoptime: ", start_time, stop_time)
        body = {
            "content": case['param']['query'],
            "senderId": case['senderId'],
            "externalSessionId": "",
            "productId": case['pid'],
            "recordId": str(uuid.uuid1()).replace('-', ''),
            "expressNumber": "",
            "command": case['param']['command'],
            "type": "1",
            "isPlaying": "false",
            "playing": {
                "content": "",
                "type": ""
            },
            "recordingFile": "",
            "startTime": starttime.strftime('%Y-%m-%d-%H-%M-%S'),
            "stopTime": stoptime.strftime('%Y-%m-%d-%H-%M-%S'),
            "asrConfidence": "1.0",
            "callNum": case['param']['callNum'],
            "sessionId": case['sessionId'],
            "outboundStartTime": "",
            "outboundEndTime": "",
            "outboundStatus": "",
            "destNumber": "",
            "hangupReason": "other",
            "cityCode": "",
            "taskId": case["outboundconfig_param"]["taskId"],
            "phoneId": case["outboundconfig_param"]["phoneId"]
        }
        if case.get("kwargs") is not None:
            body.update(case.get("kwargs"))
        logger.info("postba请求参数：{} ".format(json.dumps(body, ensure_ascii=False)))
        # print('body is {}'.format(body))
        result = requests.post(url=url, json=body)
        # print("senderId %s", case['senderId'])
        logger.info("status_code {} status_text {}".format(result.status_code, result.text))
        return result

    def postall(self, casegroup):
        """
        进行：内呼，外呼
        :param casegroup:
        :return:
        """
        # print('casegroup is {}'.format(casegroup))

        for case in casegroup:
            logger.info("case is: {}".format(json.dumps(case, ensure_ascii=False)))
            if case["outbound"] == "True" and case["param"]["query"].lower() == 'false':
                outboundr = self.postoutbound(case)
            if case["param"]["query"].lower() == "signal=hangup":
                result = self.postdmlabel(case)
                case['result'] = result
            else:
                if case["param"]["query"].lower() == 'false':
                    ctime = datetime.now()
                    start_time = ctime
                    timeago_flag = False
                    if case.get("timeago") is not None:
                        start_time = ctime - timedelta(minutes=int(case.get("timeago")))
                        endtime = ctime - timedelta(minutes=int(case.get("timeago")))
                        timeago_flag = True
                    # timeago存在则不用重新赋值endtime，否则需要给定endtime
                    if case.get("callduration") is not None:
                        if timeago_flag:
                            start_time = start_time - timedelta(seconds=int(case.get("callduration")))
                        else:
                            start_time = ctime - timedelta(seconds=int(case.get("callduration")))
                            endtime = ctime
                    result = self.postba(case, start_time=start_time, stop_time=start_time)
                elif locals().get("endtime") is not None and case["param"]["command"] == "2":
                    result = self.postba(case, start_time=start_time, stop_time=endtime)
                elif locals().get("endtime") is not None:
                    result = self.postba(case, start_time=start_time, stop_time=start_time)
                else:
                    result = self.postba(case)
                case['result'] = result.text

    def postintent(self, querys):
        print("query is ", querys)
        for query in querys:
            if query["method"] == "aichov":
                # url = self.aihivebox_url
                r = self.query_ba_aihivebox(query["query"])
                query["intent"] = r

    def postdm_all(self, cases):
        """
        多线程执行 dm 请求
        :param cases:
        :return:
        """
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.threadnum) as executor:
            to_do = []
            for case in cases:
                future = executor.submit(self.postall, case)
                to_do.append(future)

            for future in concurrent.futures.as_completed(to_do):
                # future.result()
                pass

    def handleresult_all(self, cases, wb):
        """
        多线程执行 结果对比
        :param cases:
        :return:
        """
        with concurrent.futures.ThreadPoolExecutor() as executor:
            to_do = []
            for case in cases:
                future = executor.submit(self.handledmresultt_thread, wb, case)
                to_do.append(future)

            for future in concurrent.futures.as_completed(to_do):
                # future.result()
                pass

    def compareclown_all(self, sheetnames, wb):
        """
        多线程执行 dm 请求
        :param cases:
        :return:
        """
        with concurrent.futures.ThreadPoolExecutor() as executor:
            to_do = []
            for sheetname in sheetnames:
                future = executor.submit(self.compareclown, wb, sheetname)
                to_do.append(future)

            for future in concurrent.futures.as_completed(to_do):
                # future.result()
                pass

    def postintent_all(self, querys):
        """
        多线程执行 意图 intent 请求
        :param querys:
        :return:
        """
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
            to_do = []
            for queryitem in querys:
                future = executor.submit(self.postintent(queryitem))
                to_do.append(future)

            for future in concurrent.futures.as_completed(to_do):
                # future.result()
                # future.done()
                pass

    def handledmresult(self, wb, cases):
        # sheet1 = wb[sheetname]
        for casegroup in cases:

            for case in casegroup:
                sheet1 = wb[case['sheetname']]
                print('handleresult case.yaml is{}'.format(json.dumps(case, ensure_ascii=False)))
                row = str(case['row'])
                result = case['result']
                sessionidcell = self.excelcolumnsmap.get("sessionId") + row
                # 填写sessionid
                query = case.get("param").get("query")
                if query == "false":
                    senderId = case.get("senderId")
                    sheet1[sessionidcell] = senderId
                    print("senderId is: ", senderId)

                # 应答语
                titleclown = self.excelcolumnsmap['titleclown'] + row
                # topiccode
                topiccodeclown = self.excelcolumnsmap["topiccodde"] + row
                transfernumclown = self.excelcolumnsmap["transferNum"] + row
                labelsclown = self.excelcolumnsmap["labels"] + row
                resultdict = json.loads(result)
                try:
                    labels = case.get("lablesparam")
                    if labels is None:
                        sheet1[titleclown] = resultdict['voiceText'].strip()
                        # topiccode
                        topiccde = resultdict["topicCode"]
                        # transferNum
                        transferNum = resultdict["transferNum"]
                    else:
                        sheet1[titleclown] = resultdict['best']["reply"]["voiceText"].strip()
                        # topiccode
                        topiccde = resultdict['best']["reply"]["topicCode"]
                        # 标签
                        sheet1[labelsclown] = json.dumps(resultdict.get("lablecontent"), ensure_ascii=False)
                        # transferNum
                        transferNum = resultdict['best']["reply"]["message"]["transferNum"]

                except:
                    transferNum = None
                sheet1[topiccodeclown] = topiccde
                sheet1[transfernumclown] = transferNum

    def handledmresultt_thread(self, wb, casegroup):

        for case in casegroup:
            sheet1 = wb[case['sheetname']]
            print('handleresult case.yaml is{}'.format(json.dumps(case, ensure_ascii=False)))
            row = str(case['row'])
            result = case['result']
            sessionidcell = self.excelcolumnsmap.get("sessionId") + row
            # 填写sessionid
            query = case.get("param").get("query")
            if query == "false":
                senderId = case.get("senderId")
                sheet1[sessionidcell] = senderId

            # 应答语
            titleclown = self.excelcolumnsmap['titleclown'] + row
            # topiccode
            topiccodeclown = self.excelcolumnsmap["topiccodde"] + row
            transfernumclown = self.excelcolumnsmap["transferNum"] + row
            labelsclown = self.excelcolumnsmap["labels"] + row
            order_num_clown = self.excelcolumnsmap["orderNum"] + row
            resultdict = json.loads(result)
            try:
                labels = case.get("lablesparam")
                if labels is None:
                    sheet1[titleclown] = resultdict['voiceText'].strip()
                    # topiccode
                    topiccde = resultdict["topicCode"]
                    # transferNum
                    transferNum = resultdict.get("transferNum")
                    # orderNum
                    order_num = resultdict.get("orderNum")

                else:
                    sheet1[titleclown] = resultdict['best']["reply"]["voiceText"].strip()
                    # topiccode
                    topiccde = resultdict['best']["reply"]["topicCode"]
                    # 标签
                    sheet1[labelsclown] = json.dumps(resultdict.get("lablecontent"), ensure_ascii=False)
                    # transferNum
                    transferNum = resultdict['best']["reply"]["message"]["transferNum"]

            except:
                transferNum = None

            sheet1[topiccodeclown] = topiccde
            sheet1[transfernumclown] = transferNum
            sheet1[order_num_clown].value = order_num
            serviceType = resultdict.get("serviceType")
            servicetypeclown = self.excelcolumnsmap["servicetype"] + row
            sheet1[servicetypeclown] = serviceType

    def handleintentresult(self, wb, querys):
        """
        处理意图的结果，把意图请求的list 中intent 填入 excel
        :param wb:
        :param querys:
        :return:
        """
        for queryitme in querys:
            for query in queryitme:
                sheet1 = wb[query['sheetname']]
                print('handleresult case.yaml is{}'.format(query))
                row = str(query['row'])
                result = query['intent']

                # 意图列表
                intentclown = self.excelcolumnsmap['intent'] + row
                sheet1[intentclown] = result

    def query_ba_aihivebox(self, query):
        """
        query 意图查询语料，
        :param query:
        :return: str,意图名字 "name1,name2"
        """
        dataraw = {"params": {"request": {"refText": query, "coreType": "cn.dlg.ita", "res": "aihivebox",
                                          "env": "use_slot_index=1"}}}
        data = "params" + "=" + json.dumps(dataraw["params"], ensure_ascii=False)
        print(data)
        try:
            r = requests.post(url=self.aihivebox_url, data=data.encode("utf-8"))
            if r.status_code == 200:
                param = json.loads(r.text)["result"]["semantics"]["request"]["param"]
                rstr = ''
                for k, v in param.items():
                    rstr = rstr + str(k) + ":" + str(v) + ","
                    return rstr.strip(",")
            else:
                return "意图参数问题，返回非200"
            # print(r.text)
        except:
            return "意图请求错误"

    def sessionslabels(self, endtime, starttime, sessonids, checklable, pid, env):
        # time.sleep(30)  # 等待30s保证日期都已入库es
        lables = {}
        lablecontent = {}
        # kibana=acquirekibana(query,starttime,endtime)
        # kibana = acquirekibana(self.env)

        # labeldict = (starttime, endtime, self.pid)
        labeldict = self.getlabel(starttime, endtime, pid, env)
        logger.info("sessionslabels 处理sessonids: {}".format(sessonids))
        for sessonid in sessonids:
            if labeldict.keys().__contains__(sessonid):
                pass
            else:
                time.sleep(5)
                labeldict = self.getlabel(starttime, endtime, pid, env)

        logger.info("sessionslabels 处理es内容：{}".format(json.dumps(labeldict, ensure_ascii=False)))
        for sessionid in sessonids:
            # for k, v in labeldict.items():
            if sessionid in labeldict.keys():
                for label in checklable.keys():
                    # originallabel = json.loads(labeldict.get(sessionid))
                    originallabel = labeldict.get(sessionid)
                    if label in originallabel.keys():
                        lablecontent[checklable[label]] = originallabel[label]
                    else:
                        lablecontent[checklable[label]] = "此标签不存在"

                lables[sessionid] = deepcopy(lablecontent)
                lablecontent.clear()
            else:
                lables[sessionid] = "此session未查询到标签内容"
        logger.info("sessionslabels 处理后的labels: {}".format(json.dumps(lables, ensure_ascii=False)))
        return lables

    def getlabel(self, gte, lte, pid, env):
        """

        :param gte: 开始时间
        :param lte: 结束时间
        :param pid: productid
        :return:
        """
        kibana = acquirekibana(env)
        query = [
            # {"module": "dm-engine", "productId": pid, "eventName": "contextAttrs", "operator": "是"}
            {"module": "dataclean-message", "productId": pid,
             "eventName": "http://smart-provider-boot.his.svc.cluster.local:9090/provider/api/v1/sendCallInfo",
             "operator": "是"}
        ]
        re = kibana.getlog(query, gte, lte)
        if re.status_code == 200:
            try:
                return self.parselog(re.json())
            except:
                time.sleep(15)
                re = kibana.getlog(query, gte, lte)
                if re.status_code == 200:
                    return self.parselog(re.json())
                else:
                    logger.info("es 无数据")
                    return {}
        else:
            logger.info("es 无数据")
            return {}

    def parselog(self, log):
        hits = log["responses"][0]["hits"]['hits']
        temp = {}
        for item in hits:
            sessionId = item["_source"]["sessionId"]
            offset = item["_source"]["offset"]
            message = item["_source"]["message"]
            if sessionId in temp.keys():
                message_logtime = []
                message_logtime.append(message)
                message_logtime.append(offset)
                temp[sessionId].append(deepcopy(message_logtime))
                message_logtime.clear()
            else:
                temp[sessionId] = []
                message_logtime = []
                message_logtime.append(message)
                message_logtime.append(offset)
                temp[sessionId].append(deepcopy(message_logtime))
                message_logtime.clear()
        # 排序拿出对话中最后一条标签内容
        for k, v in temp.items():
            v.sort(key=lambda x: x[1])
        # print(json.dumps(temp))
        # 截取标签值
        for k, v in temp.items():
            logger.info('处理前id {},message: {}'.format(k, v))
            last_message = v[-1][0]
            input_origin = json.loads(json.loads(last_message)["message"])["input_origin"]
            logger.info("input_origin: {}".format(json.dumps(input_origin, ensure_ascii=False)))
            reqstr = input_origin.split("req:")[1]
            callInfo = json.loads(json.loads(reqstr)["callInfo"])
            logger.info("req is : {}".format(json.dumps(callInfo, ensure_ascii=False)))
            try:
                if callInfo["fsReqInfo"]["command"] == "2":
                    temp[k] = callInfo
                else:
                    temp[k] = "此sessionid没有 hangup"
            except:
                temp[k] = "此sessionid没有 hangup"
        logger.info("es处理后数据： {}".format(json.dumps(temp, ensure_ascii=False)))
        return temp

    def postdmlabel(self, case):
        """
        从dm拿到标签
        :param case:
        :return:返回带标签的result
        """
        url = self.dmruntimeurl
        print('url is {}'.format(url))
        lablecontent = {}
        body = {
            "senderId": case["senderId"],
            "asrConfidence": 1.0,
            "productId": case['pid'],
            "query": "signal=hangUp",
            "sessionId": case['sessionId'],
            "robotId": case['pid'],
            "message": {
                "relayNum": "",
                "phone": case['param']['callNum'],
                "phoneId": "",
                "callNum": case['param']['callNum'],
                "destNumber": "",
                "taskId": ""
            }
        }
        result = requests.post(url=url, json=body)
        logger.info("request %s", body)
        logger.info("status_code %d status_text %s" % (result.status_code, result.text))
        resultdict = json.loads(result.text)
        originallabel = resultdict["best"]["reply"].get("message")
        for label in case['lablesparam'].keys():
            try:
                if label in originallabel.keys():
                    lablecontent[case['lablesparam'][label]] = originallabel.get(label)

                else:
                    lablecontent[case['lablesparam'][label]] = "此标签不存在"
            except:
                lablecontent[case['lablesparam'][label]] = "此标签不存在"

        resultdict["lablecontent"] = lablecontent
        resultstr = json.dumps(resultdict, ensure_ascii=False)
        logger.info("postdmlable resp: %s" % resultstr)
        return resultstr

    def main_data(self):
        starttime = time.perf_counter()
        wb = load_workbook(self.excelpath)
        # print(wb.sheetnames)
        # 若不给定 表单名则获取全部sheet表进行运行
        if len(self.sheetnames) == 0:
            self.sheetnames = wb.sheetnames
        # 重置字体颜色
        for sheetname in self.sheetnames:
            self.resetfont(wb=wb, sheetname=sheetname)

        cases = []
        for sheetname in self.sheetnames:
            cases.extend(self.collectcaseba(wb=wb, sheetname=sheetname, phonenumber=self.phone))
        logger.info('cases is {}'.format(cases))
        countcase = 0
        for case in cases:
            countcase=countcase+len(case)
        print("countcase is : ",countcase)
        print("groupcase is : ", len(cases))

        for i in range(10):
            self.postdm_all(cases)

    def main(self):
        starttime = time.perf_counter()
        wb = load_workbook(self.excelpath)
        # print(wb.sheetnames)
        # 若不给定 表单名则获取全部sheet表进行运行
        if len(self.sheetnames) == 0:
            self.sheetnames = wb.sheetnames
        # 重置字体颜色
        for sheetname in self.sheetnames:
            self.resetfont(wb=wb, sheetname=sheetname)

        cases = []
        querys = []
        for sheetname in self.sheetnames:
            cases.extend(self.collectcaseba(wb=wb, sheetname=sheetname, phonenumber=self.phone))
            querys.extend(self.collectquery(wb=wb, sheetname=sheetname))
        logger.info('cases is {}'.format(cases))
        logger.info("querys is {}".format(querys))
        # countcase=0
        # for case in cases:
        #     countcase=countcase+len(case)
        # print("countcase is : ",countcase)
        self.postdm_all(cases)
        # self.handledmresult(wb, cases)
        print("cases is :", json.dumps(cases, ensure_ascii=False))
        self.handleresult_all(cases=cases, wb=wb)
        if len(querys) > 0:
            self.postintent_all(querys)
            self.handleintentresult(wb, querys)
        # 循环标记结果
        # for sheetname in self.sheetnames:
        #     self.compareclown(wb, sheetname)
        self.compareclown_all(sheetnames=self.sheetnames, wb=wb)
        wb.save(self.excelpath)

        # pass

        endtime = time.perf_counter()
        print('{} time cost {}'.format(self.excelpath, (endtime - starttime)))
        with open("test_reuslt", "a", encoding="utf-8") as result_file:
            if len(self.sheetname_wrong) > 0:
                result_file.write(
                    "运行结束时间：  " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "   存在失败case的用例表  " + str(
                        self.sheetname_wrong))
            else:
                result_file.write(
                    "运行结束时间：  " + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "   执行通过用例表  " + str(
                        self.sheetnames))
            result_file.write("\n")
            print(self.sheetname_wrong)


if __name__ == '__main__':
    phone = '1025'
    # mobilephone = '18513583960'
    env = 'test'
    excelpath = r'test.xlsx'
    sheetnames = ["Sheet1", "Sheet2"]
    dm = dialogtxt(env, phone, excelpath, sheetnames)
    # # print(dm.main())
    # dm.main()
    # dm.query_ba_aihivebox()
    num = set()
    for i in range(10001):
        # print(i)
        phone = "185" + str(dm.gettailnum() + dm.gettailnum())
        num.add(phone)
    print(len(num))
    # print(num)
